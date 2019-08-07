import psycopg2
import openpyxl
from openpyxl.styles import Font


def export_to_excel(connection, query_string, headings, filepath):

    """
    Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.

    Arguments:
    connection - an open psycopg2 (this function does not close the connection)
    query_string - SQL to get data
    headings - list of strings to use as column headings
    filepath - path and filename of the Excel file

    psycopg2 and file handling errors bubble up to calling code.
    """

    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    cursor.close()

    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()

    sheet.row_dimensions[1].font = Font(bold = True)

    # Spreadsheet row and column indexes start at 1
    # so we use "start = 1" in enumerate so
    # we don't need to add 1 to the indexes.
    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading

    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value

    wb.save(filepath)
